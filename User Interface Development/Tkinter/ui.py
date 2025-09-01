import tkinter as tk
from tkinter import ttk, messagebox
from user_functions import press_browse_file, press_save_file, press_run, press_clear_color
from openpyxl import load_workbook, Workbook, worksheet
import re
import os
import threading


ref_wb = Workbook()
ref_ws = ref_wb.active
target_wb = Workbook()
target_ws = target_wb.active

class MainApp:
    def __init__(self, root):
        self.root = root
        self.root.title("WELL REPLACEMENT FORM")

        # Reference Well Section
        self.ref_frame = ttk.LabelFrame(root, text="REFERENCE WELL")
        self.ref_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        # Progress bar update
        self.ref_loading = ttk.Progressbar(self.ref_frame, mode="indeterminate", length=20)
        self.ref_loading.grid(row=2, column=2, padx=5, pady=5)
        self.ref_loading.grid_remove()  # Hide initially

        # Reference file path label
        self.ref_file_path = ttk.Label(self.ref_frame, text="File Path:")
        self.ref_file_path.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        # Reference file path text box
        self.ref_entry = ttk.Entry(self.ref_frame, width=40)
        self.ref_entry.grid(row=0, column=1, padx=5, pady=5)

        # Create a combo box for reference sheet selection
        ttk.Label(self.ref_frame, text="Worksheet:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.ref_worksheet_cbo = ttk.Combobox(self.ref_frame, values=["Sheet1", "Sheet2", "Sheet3"])
        self.ref_worksheet_cbo.grid(row=1, column=1, padx=5, pady=5)

        # Create a combo box for reference well name selection
        ttk.Label(self.ref_frame, text="Well Name:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.ref_well_name_cbo = ttk.Combobox(self.ref_frame, values=["Well1", "Well2", "Well3"])
        self.ref_well_name_cbo.grid(row=2, column=1, padx=5, pady=5)

        # Reference file path button
        self.ref_file_button = ttk.Button(self.ref_frame, text="Browse File", 
                            command=lambda: press_browse_file(self.ref_entry, self.ref_worksheet_cbo,
                                                            self.ref_well_name_cbo, 
                                                            input_mode="reference"))
        if self.ref_entry.get():
            print(self.ref_entry.get())
        self.ref_file_button.grid(row=0, column=2, padx=5, pady=5)

        # Target Well Section
        self.target_frame = ttk.LabelFrame(root, text="TARGET WELL")
        self.target_frame.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

        ttk.Label(self.target_frame, text="File Path:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.target_entry = ttk.Entry(self.target_frame, width=40)
        self.target_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(self.target_frame, text="Browse File", 
                command=lambda: press_browse_file(self.target_entry, self.target_worksheet_cbo, self.target_well_name_cbo, input_mode="target" )).grid(row=0, column=2, padx=5, pady=5)

        # Create a combo box for target sheet selection
        ttk.Label(self.target_frame, text="Worksheet:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.target_worksheet_cbo = ttk.Combobox(self.target_frame, values=["Sheet1", "Sheet2", "Sheet3"])
        self.target_worksheet_cbo.grid(row=1, column=1, padx=5, pady=5)

        # Create a combo box for target well selection
        ttk.Label(self.target_frame, text="Well Name:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.target_well_name_cbo = ttk.Combobox(self.target_frame, values=["Well1", "Well2", "Well3"])
        self.target_well_name_cbo.grid(row=2, column=1, padx=5, pady=5)

        # # File Saving Section
        self.save_frame = ttk.LabelFrame(root, text="SAVING LOCATION")
        self.save_frame.grid(row=2, column=0, padx=10, pady=10, sticky="ew")

        ttk.Label(self.save_frame, text="File Path:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.save_entry = ttk.Entry(self.save_frame, width=40)
        self.save_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(self.save_frame, text="Browse File", 
                command=lambda: press_save_file(self.save_entry)).grid(row=0, column=2, padx=5, pady=5)

        # Options Section
        self.options_frame = ttk.LabelFrame(root, text="OPTIONS")
        self.options_frame.grid(row=3, column=0, padx=10, pady=10, sticky="ew")
        check_var = {}
        options = ["Replace with D125", "Replace WELL name", "Replace BDC", "Align Address", "No Coloring"]
        for i, option in enumerate(options):
            var = tk.BooleanVar()
            check_var[option] = var
            ttk.Checkbutton(self.options_frame, 
                            text=option, 
                            variable=var,
                            onvalue=True,
                            offvalue=False).grid(row=i, column=0, padx=5, pady=5, sticky="w")

        # Buttons at the bottom
        self.button_frame = ttk.Frame(root)
        self.button_frame.grid(row=4, column=0, padx=10, pady=10, sticky="ew")

        ttk.Button(self.button_frame, text="Quit", command=root.quit).grid(row=0, column=0, padx=5, pady=5)
        ttk.Button(self.button_frame, text="Clear Contents", 
                command=lambda: press_clear_color()).grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(self.button_frame, text="Run", 
                    command=lambda: press_run(
                        check_var,
                        ref_wb,
                        ref_ws,
                        target_wb,
                        target_ws,
                        self.ref_well_name_cbo.get(),
                        self.target_well_name_cbo.get(),
                        self.save_entry.get()
                        )
                    ).grid(row=0, column=2, padx=5, pady=5)
        
        # Binding events
        self.ref_worksheet_cbo.bind('<<ComboboxSelected>>', lambda e: self.on_combobox_change(e, mode="reference"))
        self.target_worksheet_cbo.bind('<<ComboboxSelected>>', lambda e: self.on_combobox_change(e, mode="target"))


    def on_combobox_change(self, _event, mode):
        def worker():
            regEx_Mobus = "Modbus"
            well_col = 18
            regEx_BDC = re.compile(r".*BDC.*WELL.*")
            well_names = set()

            sheet_selection = self.ref_worksheet_cbo.get() if mode == "reference" else self.target_worksheet_cbo.get()
            file_path_sel = self.ref_entry.get() if mode == "reference" else self.target_entry.get() 

            self.root.after(0, lambda: self.ref_loading.grid() if mode == "reference" else None)
            self.root.after(0, lambda: self.ref_loading.start() if mode == "reference" else None)

            valid_extensions = ('.xlsx', '.xlsm', '.xltx', '.xltm')
            if not file_path_sel or not os.path.isfile(file_path_sel) or not file_path_sel.lower().endswith(valid_extensions):
                self.root.after(0, lambda: messagebox.showerror("File Error", "Please select a valid Excel file (.xlsx, .xlsm, .xltx, .xltm)."))
                self.root.after(0, lambda: self.ref_well_name_cbo['values'] if mode == "reference" else self.target_well_name_cbo['values'] == [])
                return

            if re.findall(regEx_Mobus, sheet_selection, flags=re.IGNORECASE):
                try:
                    wb = load_workbook(file_path_sel)
                    ws = wb[sheet_selection]
                    for row in ws.iter_rows(min_row=1, values_only=True):
                        cell_value = row[well_col] if len(row) > well_col else None
                        if cell_value and regEx_BDC.search(str(cell_value)):
                            well_names.add(cell_value)
                    well_names.add("All wells")
                    wb.close()
                    # Update GUI in main thread
                    def update_gui():
                        if mode == "reference":
                            self.ref_well_name_cbo['values'] = sorted(well_names)
                            global ref_wb, ref_ws
                            ref_wb = wb
                            ref_ws = ws
                        else:
                            self.target_well_name_cbo['values'] = sorted(well_names)
                            global target_wb, target_ws
                            target_wb = wb
                            target_ws = ws
                    self.root.after(0, update_gui)
                except Exception as e:
                    self.root.after(0, lambda: messagebox.showerror("Excel Error", f"Failed to open Excel file or sheet: {e}"))
                    self.root.after(0, lambda: self.ref_well_name_cbo['values'] if mode == "reference" else self.target_well_name_cbo['values'] == [])
                finally:
                    self.root.after(0, lambda: self.ref_loading.stop() if mode == "reference" else None)
                    self.root.after(0, lambda: self.ref_loading.grid_remove() if mode == "reference" else None)
            else:
                self.root.after(0, lambda: messagebox.showerror("Error Status", "Please select the Modbus sheet"))
                self.root.after(0, lambda: self.ref_well_name_cbo['values'] if mode == "reference" else self.target_well_name_cbo['values'] == [])

        threading.Thread(target=worker).start()