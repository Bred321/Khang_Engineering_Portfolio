from openpyxl import load_workbook
import re


if __name__ == '__main__':
    result_well_str = "BDC2_WELL2"
    input_well_str = "BDC1A_WELL3"

    result_well = result_well_str.split("_")[-1]
    change_well = input_well_str.split("_")[-1]
    print(result_well)
    print(change_well)
    well_regEx = rf"{change_well}"
    print(re.search(well_regEx, input_well_str, flags=re.IGNORECASE))
    new_str = re.sub(well_regEx, result_well, input_well_str)
    print(new_str)
    # wb = load_workbook("Uni_XT_P1_MCS-ICSS MODBUS IO Schedule-REV 00.xlsx", read_only=True, data_only=True)
    # modbus_ws = wb['MODBUS']
    # well_name = set()
    # well_col = 18
    # regEx_BDC=".*BDC.*WELL.*"
    # print([ws.title for ws in wb.worksheets])
    # # Print column names (headers in the first row)
    # print(type(modbus_ws))
    # for row in modbus_ws.iter_rows(min_row=0):
    #     well_name.add(row[well_col].value if re.search(regEx_BDC, row[well_col].value) else "")
    # # print(headers)
    # print(well_name)
    # wb.close()
